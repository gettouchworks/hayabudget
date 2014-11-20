#coding=utf-8
from openpyxl import load_workbook
from openpyxl import Workbook
from sqlitedb import DB
import scheme
from scheme import *
import sys
import string

reload(sys)
sys.setdefaultencoding('utf8')

def budget_summary():
	# print querys
	filename = "dox/模板.xlsx"
	# wb = Workbook() #load_workbook()
	# ws = wb.create_sheet()
	# ws.title = u"全国汇总表"
	# ws.cell(row = rx,column = 1).value
	cols = string.uppercase[:16]
	conn = DB("haya.db3")
	op1 = "sum(a.[%s]) 工资" % "]+a.[".join(querys["工资"])
	op11 = "sum(a.[%s]) 奖金" % "]+[".join(querys["奖金"])
	op12 = "sum(a.[%s]) 补偿金" % "]+[".join(querys["补偿"])
	op2 = "sum(a.[%s]) 社保" % "]+a.[".join(querys["社保"])
	op3 = "sum(a.[%s]) 住房" % "]+a.[".join(querys["住房"])
	op4 = "sum(a.[%s]) 福利" % "]+a.[".join(querys["福利"])
	op5 = "sum(a.[%s]) 教育" % "]+a.[".join(querys["教育"])
	op6 = "sum(a.[%s]) 工会" % "]+a.[".join(querys["工会"])
	op7 = "sum(a.[%s]) 劳保" % "]+a.[".join(querys["劳保"])

	sql = '''select 
		a.实际工作地, 
		a.费用类别, 
		a.费用分类, 
		a.计薪方式, 
		a.各体系负责人, 
		a.考核单元负责人, 
		a.一级部门, 
		a.二级部门, 
		sum(人数) 人数, 
		%s, %s, %s, %s, %s, %s, %s, %s, %s 
		,b.社保合规总额, b.住房合规总额, b.社保稽查补缴
		from haya_budget as a
		left join (
			select 
				[实际工作地],
				[二级部门], 
				sum([社保合规总额]) 社保合规总额, 
				sum([住房合规总额]) 住房合规总额,
				sum([社保稽查补缴]) 社保稽查补缴 
			from haya_addons group by [实际工作地],[二级部门]) as b
		  on a.实际工作地 = b.实际工作地 and a.二级部门=b.二级部门
		 group by a.实际工作地, a.二级部门
		''' % (op1, op11, op12, op2, op3, op4, op5, op6, op7)
	print sql
	rs = conn.getData(sql)
	conn.closeDb()
	return rs
	line = 4
	for data in rs:
		print data["实际工作地"], data["二级部门"], data['工资']
		for c in range(len(data)):
		# print data["工资发放地"], data["二级部门"], data['工资']
			ws.cell("%s%s" % (cols[c],line)).value = '%s' % (data[c])
		line = line+1
	conn.closeDb()
	wb.save(filename = "text1.xlsx")

def budget_local(type=1):
	# print querys
	# wb = Workbook() #load_workbook()
	# ws = wb.create_sheet()
	# ws.title = u"按属地部门分类汇总1"
	# ws.cell(row = rx,column = 1).value
	location = "工资发放地" if type==1 else "实际工作地"

	cols = string.uppercase[:16]
	conn = DB("haya.db3")
	op1 = "sum([%s]) 工资" % "]+[".join(querys["工资"])
	op2 = "sum([%s]) 奖金" % "]+[".join(querys["奖金"])
	op3 = "sum([%s]) 补偿金" % "]+[".join(querys["补偿"])
	op4 = ("sum([%s]) 社保合计," % "]+[".join(querys["社保"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["社保"]])) #"sum([%s])" % "]),sum([".join(querys["社保"])
	op5 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["住房"]]))
	op6 = ("sum([%s]) 餐补合计," % "]+[".join(querys["餐补"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["餐补"]]))
	op7 = ("sum([%s]) 交通合计," % "]+[".join(querys["交通"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["交通"]]))
	op8 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["通信"]]))
	op9 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["探亲"]]))
	op10 = ("sum([%s]) 节日合计," % "]+[".join(querys["节日"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["节日"]]))
	op11 = ("sum([%s]) 医疗合计," % "]+[".join(querys["医疗"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["医疗"]]))
	op12 = ("sum([%s]) 固定活动合计," % "]+[".join(querys["固定活动"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["固定活动"]]))
	op13 = ("sum([%s]) 其他活动合计," % "]+[".join(querys["其他活动"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["其他活动"]]))
	op14 = ("sum([%s]) 其他福利合计," % "]+[".join(querys["其他福利"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["其他福利"]]))
	op15 = ("sum([%s]) 教育合计," % "]+[".join(querys["教育"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["教育"]]))
	op16 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["工会"]]))
	op17 = ("sum([%s]) 劳保合计," % "]+[".join(querys["劳保"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["劳保"]]))

	# querys["社保"]
	# querys["住房"]
	# querys['餐补']
	# querys['交通']
	# querys['通信']
	# querys['探亲']
	# querys['节日']
	# querys['医疗']
	# querys['固定活动']
	# querys['其他活动']
	# querys['其他福利']
	# querys['教育'] 
	# querys['工会'] 
	# querys['劳保'] 
	# querys["福利"] 

	sql = '''select 
		%s, 
		费用类别, 
		费用分类, 
		计薪方式,
		预算单位, 
		各体系负责人, 
		考核单元负责人, 
		一级部门, 
		二级部门, 
		sum(人数) 人数, 
		%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
		from haya_budget group by %s, 二级部门
		''' % (location, op1, op2, op3, op4, op5, op6, op7, op8, op9, op10, op11, op12, op13, op14,op15, op16, op17, location)
	print sql
	rs = conn.getData(sql)
	conn.closeDb()
	return rs
	line = 4
	for data in rs:
		print data["工资发放地"], data["二级部门"], data['工资']
		for c in range(len(data)):
		# print data["工资发放地"], data["二级部门"], data['工资']
			# ws.cell("%s%s" % (cols[c],line)).value = '%s' % (data[c])
			ws.cell(column=c,row=line).value = '%s' % (data[c])
		line = line+1
	conn.closeDb()
	wb.save(filename = "text2.xlsx")

def budget_state(type=1):
	# print querys
	# wb = Workbook() #load_workbook()
	# ws = wb.create_sheet()
	# ws.title = u"按属地部门分类汇总1"
	# ws.cell(row = rx,column = 1).value
	location = "工资发放地"

	cols = string.uppercase[:16]
	conn = DB("haya.db3")
	op1 = "sum([%s]) 工资" % "]+[".join(querys["工资"])
	op2 = "sum([%s]) 奖金" % "]+[".join(querys["奖金"])
	op3 = "sum([%s]) 补偿金" % "]+[".join(querys["补偿"])
	op4 = ("sum([%s]) 社保合计," % "]+[".join(querys["社保"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["社保"]])) #"sum([%s])" % "]),sum([".join(querys["社保"])
	op5 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["住房"]]))
	op6 = ("sum([%s]) 餐补合计," % "]+[".join(querys["餐补"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["餐补"]]))
	op7 = ("sum([%s]) 交通合计," % "]+[".join(querys["交通"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["交通"]]))
	op8 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["通信"]]))
	op9 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["探亲"]]))
	op10 = ("sum([%s]) 节日合计," % "]+[".join(querys["节日"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["节日"]]))
	op11 = ("sum([%s]) 医疗合计," % "]+[".join(querys["医疗"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["医疗"]]))
	op12 = ("sum([%s]) 固定活动合计," % "]+[".join(querys["固定活动"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["固定活动"]]))
	op13 = ("sum([%s]) 其他活动合计," % "]+[".join(querys["其他活动"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["其他活动"]]))
	op14 = ("sum([%s]) 其他福利合计," % "]+[".join(querys["其他福利"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["其他福利"]]))
	op15 = ("sum([%s]) 教育合计," % "]+[".join(querys["教育"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["教育"]]))
	op16 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["工会"]]))
	op17 = ("sum([%s]) 劳保合计," % "]+[".join(querys["劳保"])) + (",".join(["sum([%s]) %s" % (item,item) for item in querys["劳保"]]))


	sql = '''select 
		%s, 
		费用类别, 
		费用分类, 
		预算单位, 
		各体系负责人, 
		考核单元负责人, 
		状态,
		一级部门, 
		二级部门, 
		sum(人数) 人数, 
		%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
		from haya_budget group by %s, 状态, 二级部门
		''' % (location, op1, op2, op3, op4, op5, op6, op7, op8, op9, op10, op11, op12, op13, op14,op15, op16, op17, location)
	print sql
	rs = conn.getData(sql)
	conn.closeDb()
	return rs

def budget_all():

	conn = DB("haya.db3")
	op1 = (",".join(["sum([%s]) '%s'" % (item,item) for item in querys["工资"]])) #"sum([%s]) 工资" % "]+[".join(querys["工资"])
	op2 = "sum([%s]) 奖金" % "]+[".join(querys["奖金"])
	op3 = "sum([%s]) 补偿金" % "]+[".join(querys["补偿"])
	op4 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["社保"]])) #"sum([%s])" % "]),sum([".join(querys["社保"])
	op5 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["住房"]]))
	op6 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["餐补"]]))
	op7 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["交通"]]))
	op8 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["通信"]]))
	op9 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["探亲"]]))
	op10 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["节日"]]))
	op11 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["医疗"]]))
	op12 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["固定活动"]]))
	op13 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["其他活动"]]))
	op14 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["其他福利"]]))
	op15 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["教育"]]))
	op16 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["工会"]]))
	op17 = (",".join(["sum([%s]) %s" % (item,item) for item in querys["劳保"]]))

	sql = '''select 
		公司名称, 
		工资发放地,
		实际工作地,
		费用类别, 
		费用分类, 
		预算单位,
		计薪方式,
		各体系负责人, 
		考核单元负责人,
		状态, 
		一级部门, 
		二级部门, 
		sum(人数) 人数,
		员工编号,
		姓名,
		性别,
		岗位名称,
		职级,
		%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
		from haya_budget group by 二级部门, 员工编号, 姓名
		''' % (op1, op2, op3, op4, op5, op6, op7, op8, op9, op10, op11, op12, op13, op14,op15, op16, op17)
	print sql
	rs = conn.getData(sql)
	conn.closeDb()
	return rs

def budget_staff():
	location = "实际工作地"
	cols = string.uppercase[:16]
	conn = DB("haya.db3")

	sql = '''select 
		月份,
		%s, 
		一级部门, 
		二级部门, 
		sum(人数) 人数
		from haya_budget group by %s, 二级部门, 月份
		''' % (location, location)
	print sql
	rs = conn.getData(sql)
	conn.closeDb()
	items = {}
	for data in rs:
		key = "%s_%s_%s" % (data["实际工作地"], data["一级部门"], data["二级部门"])
		if key in items:
			items[key][data["月份"]] = data["人数"]
		else:
			items[key] = {u"实际工作地":data["实际工作地"],u"一级部门":data["一级部门"],u"二级部门":data["二级部门"],data["月份"]:data["人数"]}
	return items.values()
	# return rs

def init_table():
	print scheme.fields
	sql = """create table haya_budget (
		id integer auto_increment, filename varchar(100), mouth varchar(100), 
		'公司名称'  varchar(100),
		'属地1'  varchar(100),
		'属地2'  varchar(100),
		'费用类别'  varchar(100),
		'费用分类'  varchar(100),
		'预算单位'  varchar(100),
		'计薪方式'  varchar(100),
		'各体系负责人'  varchar(100),
		'考核单元负责人'  varchar(100),
		'状态'  varchar(100),
		'一级部门'  varchar(100),
		'二级部门'  varchar(100),
		'2015年人数'  varchar(100),
		'员工编号'  varchar(100),
		'姓名'  varchar(100),
		'性别'  varchar(100),
		'岗位名称'  varchar(100),
		'职级'  varchar(100),
		'在职月数' numeric,
		'基本工资' numeric,
		'绩效工资' numeric,
		'岗位工资' numeric,
		'固定加班工资' numeric,
		'其他工资' numeric,
		'工龄工资' numeric,
		'房屋津贴' numeric,
		'误工费' numeric,
		'市场误餐费' numeric,
		'午餐补贴' numeric,
		'其他津贴' numeric,
		'项目制奖金' numeric,
		'成本节约考核奖' numeric,
		'6S奖' numeric,
		'创新奖' numeric,
		'特殊类人员奖金' numeric,
		'其他奖金' numeric,
		'绩效奖金' numeric,
		'奖金池' numeric,
		'平日加班工资' numeric,
		'周末加班工资' numeric,
		'法定假日加班工资' numeric,
		'用于日常现金报销的奖励' numeric,
		'经济补偿金' numeric,
		'社保基数' numeric,
		'养老保险' numeric,
		'失业保险' numeric,
		'医疗保险' numeric,
		'工伤保险' numeric,
		'生育保险' numeric,
		'特殊保险' numeric,
		'住房公积金数据' numeric,
		'住房公积金' numeric,
		'正常午餐补助' numeric,
		'加班餐补' numeric,
		'定额交通报销' numeric,
		'定额汽油' numeric,
		'定额通讯补助' numeric,
		'探亲路费' numeric,
		'妇女节礼品费' numeric,
		'中秋节日费' numeric,
		'体检费' numeric,
		'医疗费' numeric,
		'员工活动经费' numeric,
		'生日费' numeric,
		'年会费' numeric,
		'年度评优费用' numeric,
		'庆典' numeric,
		'理想家平台费' numeric,
		'羽毛球' numeric,
		'运动会' numeric,
		'比赛竞赛' numeric,
		'其他活动' numeric,
		'存档费' numeric,
		'结婚喜金' numeric,
		'生育礼金' numeric,
		'管理人员房租' numeric,
		'其他补助' numeric,
		'职称考试评定' numeric,
		'企业培训' numeric,
		'职工教育经费其他' numeric,
		'工会经费' numeric,
		'劳保用品' numeric,
		'工服' numeric,
		'防暑降温' numeric,
		'劳动保护其他' numeric
		)"""
	print sql
	"""CREATE TABLE "haya_budget" ("id" integer PRIMARY KEY ,"文件" varchar (100),"月份" varchar (100),"公司名称" varchar (100),"工资发放地" varchar (100) DEFAULT (null) ,"实际工作地" varchar (100) DEFAULT (null) ,"费用类别" varchar (100),"费用分类" varchar (100),"预算单位" varchar (100),"计薪方式" varchar (100),"各体系负责人" varchar (100),"考核单元负责人" varchar (100),"状态" varchar (100),"一级部门" varchar (100),"二级部门" varchar (100),"人数" varchar (100) DEFAULT (null) ,"员工编号" varchar (100),"姓名" varchar (100),"性别" varchar (100),"岗位名称" varchar (100),"职级" varchar (100),"在职月数" numeric,"基本工资" numeric,"绩效工资" numeric,"岗位工资" numeric,"固定加班工资" numeric,"其他工资" numeric,"工龄工资" numeric,"房屋津贴" numeric,"误工费" numeric,"市场误餐费" numeric,"午餐补贴" numeric,"其他津贴" numeric,"项目制奖金" numeric,"成本节约考核奖" numeric,"6S奖" numeric,"创新奖" numeric,"特殊类人员奖金" numeric,"其他奖金" numeric,"绩效奖金" numeric,"奖金池" numeric,"平日加班工资" numeric,"周末加班工资" numeric,"法定假日加班工资" numeric,"用于日常现金报销的奖励" numeric,"经济补偿金" numeric,"社保基数" numeric,"养老保险" numeric,"失业保险" numeric,"医疗保险" numeric,"工伤保险" numeric,"生育保险" numeric,"特殊保险" numeric,"住房公积金数据" numeric,"住房公积金" numeric,"正常午餐补助" numeric,"加班餐补" numeric,"定额交通报销" numeric,"定额汽油" numeric,"定额通讯补助" numeric,"探亲路费" numeric,"妇女节礼品费" numeric,"中秋节日费" numeric,"体检费" numeric,"医疗费" numeric,"员工活动经费" numeric,"生日费" numeric,"年会费" numeric,"年度评优费用" numeric,"庆典" numeric,"理想家平台费" numeric,"羽毛球" numeric,"运动会" numeric,"比赛竞赛" numeric,"其他活动" numeric,"存档费" numeric,"结婚喜金" numeric,"生育礼金" numeric,"管理人员房租" numeric,"其他补助" numeric,"职称考试评定" numeric,"企业培训" numeric,"职工教育经费其他" numeric,"工会经费" numeric,"劳保用品" numeric,"工服" numeric,"防暑降温" numeric,"劳动保护其他" numeric)"""
	conn = DB("haya.db3")
	conn.runSql(sql)

def init_addons():
	"""CREATE TABLE "haya_addons" ("id" INTEGER PRIMARY KEY  AUTOINCREMENT  NOT NULL , "月份" VARCHAR, "公司名称" VARCHAR, "工资发放地" VARCHAR, "实际工作地" VARCHAR, "费用类别" VARCHAR, "费用分类" VARCHAR, "预算单位" VARCHAR, "计薪方式" VARCHAR, "各体系负责人" VARCHAR, "考核单元负责人" VARCHAR, "状态" VARCHAR, "一级部门" VARCHAR, "二级部门" VARCHAR, "人数" VARCHAR, "员工编号" VARCHAR, "姓名" VARCHAR, "性别" VARCHAR, "岗位名称" VARCHAR, "职级" VARCHAR, "在职月数" VARCHAR, "月社保基数增加额" NUMERIC, "社保基数合规月数" NUMERIC, "月养老增加额" NUMERIC, "月工伤增加额" NUMERIC, "月失业增加额" NUMERIC, "月生育增加额" NUMERIC, "月医疗增加额" NUMERIC, "社保合规总额" NUMERIC, "住房基数增加额" NUMERIC, "合规月数" DATETIME, "月住房公积金增加额" NUMERIC, "住房合规总额" NUMERIC, "社保稽查补缴" NUMERIC)"""

def create_table(tbname):
	sql = "SELECT name FROM sqlite_master WHERE type='table' AND name='%s'" % tbname

	conn = DB("haya.db3")
	rs = conn.getLine(sql)
	print rs[0]
	if not rs:
		sql = "create table '%s' (id integer primary key,pid integer,name varchar(10) UNIQUE)" % tbname
		conn.runSql(sql)

def load_file(filename=''):
	if not filename:
		return false
	print filename
	conn = DB("haya.db3")

	wb = load_workbook(filename = filename)

	sheets = scheme.monthList
	ufields = scheme.fields
	del ufields[0]
	for x in sheets:
		print x
		ws = wb.get_sheet_by_name(name = x)
		if not ws:
			continue
		print "Work Sheet Titile:",ws.title 
		rows = ws.rows
		for r in xrange(8, len(rows)):
			row = rows[r]
			if not row[1].value:
				break
			# for c in xrange(1, len(ufields)+1):
			# 	print row[c].value
			data = [str(d.value) if d.value else "0" for d in list(row[1:len(ufields)+1])]
			sql = "insert into haya_budget values (null, '', '%s', '%s')" % (x, "','".join(data))
			print sql
			# break
			conn.runSql(sql)
		# break
		# cc = ws.get_highest_column()
		# rc = ws.get_highest_row()
		# print cc, rc
		# # for row in ws.iter_rows(row_offset=8):
		
		# for rx in xrange(8, rc):
		# 	if not ws.cell(row = rx,column = 1).value:
		# 		break
		# 	row = []
		# 	for cx in xrange(1, len(ufields)+1):
		# 		# print cx
		# 		v =  ws.cell(row = rx,column = cx).value
		# 		# print v
		# 		if not v:
		# 			v = 0
		# 		v = str(v)
		# 		row.append(v)
		# 		# print x, rx, scheme.fields[cx-1], ws.cell(row = rx,column = cx).value
		# 	print len(row)
		# 	data = "','".join(row)
		# 	# print data
			
# init_table()

def dumpfile():
	# load_file(u"/Users/gettouch/Projects/pythons/excel/dox/2015年北京总部人力预算.xlsx");

	# rs = budget_staff()
	
	# for item in rs:
	# 	print item
	# sys.exit()
	# budget_summary()
	datalist = [
		budget_summary(),
		# budget_local(1),
		budget_local(2),
		budget_all(),
		budget_state(),
		# budget_staff(),
		]
	sheetlist = [u"全国汇总表", u"实际工作地", u"全国明细汇总表", u"人员状态", u"人员统计"]
	# result = budget_all()
	wb = Workbook(guess_types=True) #load_workbook()
	# for i in range(len(datalist)):
	for i, result in enumerate(datalist):
		# result = datalist[i]
		ws = wb.create_sheet()
		ws.title = sheetlist[i]
		line = 4
		keys = result[0].keys()
		# for c in range(len(keys)):
		for c, item in enumerate(keys):
			ws.cell(column=c+1,row=3).value = '%s' % (item)
		# for data in result:
		for line, data in enumerate(result):
			# print data.keys()
			# print data["工资发放地"], data["二级部门"], data['工资']
			# for c in range(len(data)):
			for c, item in enumerate(data):
			# print data["工资发放地"], data["二级部门"], data['工资']
				# ws.cell("%s%s" % (cols[c],line)).value = '%s' % (data[c])
				ws.cell(column=c+1,row=line+4).value = '%s' % (item if item else "0")
			# line = line+1
	staffs = budget_staff()
	ws = wb.create_sheet()
	ws.title = u"人员统计"
	keys = staffs[0].keys()
	base_keys = [u"实际工作地", u"一级部门", u"二级部门", u"1月", u"2月", u"3月", u"4月", u"5月", u"6月", u"7月", u"8月", u"9月", u"10月", u"11月", u"12月"]
	for c, item in enumerate(base_keys):
		ws.cell(column=c+1,row=3).value = '%s' % (item)
	for line, data in enumerate(staffs):
		print data
		for c, item in enumerate(base_keys):
			print item
			ws.cell(column=c+1,row=line+4).value = '%s' % (data[item])

	wb.save(filename = u"预算汇总.xlsx")

if __name__ == '__main__':
	print budget_summary()
	# dumpfile()
# wb = load_workbook(filename = '人力预算.xlsx')
# # print "Worksheet range(s):", wb.get_named_ranges() 
# # sheets = wb.get_sheet_names()
# # print sheets
# sheets = ['0月', '1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
# for x in sheets:
# 	print x
# 	ws = wb.get_sheet_by_name(name = x)
# 	if not ws:
# 		continue
# 	print "Work Sheet Titile:",ws.title 
# 	cc = ws.get_highest_column()
# 	rc = ws.get_highest_row()
# 	print x, cc, rc
	# for rx in xrange(8, rc):   
	# 	for cx in xrange(cc):
	# 		print x, rx, cx, ws.cell(row = rx,column = cx).value
	      
	     
# sys.exit()

# ws = wb.get_sheet_by_name(name = u'工作表1')

# # print sheet_ranges
# print "Work Sheet Titile:",ws.title 
# print ws.get_highest_column()
# print ws.get_highest_row()



# data_dic = {}
# for rx in range(ws.get_highest_row()):  
      
#     temp_list = []  
#     pid = ws.cell(row = rx,column = 0).value  
#     w1 = ws.cell(row = rx,column = 1).value  
#     w2 = ws.cell(row = rx,column = 2).value  
#     w3 = ws.cell(row = rx,column = 3).value  
#     w4 = ws.cell(row = rx,column = 4).value  
#     temp_list = [w1,w2,w3,w4]  
     
#     data_dic[pid] = temp_list  
  
# print data_dic
#打印字典数据个数  
# print 'Total:%d' %len(data_dic)
# print ws.cell('A1').value # D18